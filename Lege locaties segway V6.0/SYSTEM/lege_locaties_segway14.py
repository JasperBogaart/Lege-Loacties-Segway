import os
import sys
import tkinter as tk
from collections import Counter
from pathlib import Path
from tkinter import filedialog, messagebox, ttk


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


# Vaste teksten en locatiegroepen die de tool op meerdere plekken gebruikt.
ALLE_GANGEN = "Alle Gangen"
EMPTY_STATUS = "E - Empty"
FULL_STATUS = "F - Full"
GELDIGE_LOCATIONS = ("A", "B", "C")
VERPLICHTE_KOLOMMEN = ("Location", "Location Status", "Enabled?", "Aisle", "Height", "Sto_Zone_Cod")

# Nieuwe WMS-exporten kunnen andere kolomnamen hebben. Intern blijft de oude naam gebruikt.
KOLOM_ALIASSEN = {
    "Sto_Zone_Cod": ("Storage Zone Code",),
}

# Locatienummers die niet in de normale lege-locaties-lijst horen.
UITGEZONDERDE_LOCATIONS = {
    "A": set(range(9, 211, 8)) | set(range(10, 211, 8)),
    "BBX": set(range(3, 80, 4)),
    "BC": set(range(5, 199, 8)) | set(range(6, 199, 8)),
}

# Zoekwoorden uit de storage zone worden omgezet naar leesbare family types.
FAMILY_KEYWORDS = ("big", "med-lng", "med-shrt", "sml-lng", "sml-shrt", "ground")
FAMILY_LABELS = ("BIG", "MED-LNG", "MED-SHRT", "SML-LNG", "SML-SHRT", "GROUND")
FAMILY_KLEUREN = {
    "BIG": "D9EAD3",
    "MED-LNG": "CFE2F3",
    "MED-SHRT": "EAD1DC",
    "SML-LNG": "FFF2CC",
    "SML-SHRT": "F4CCCC",
    "GROUND": "E6E0EC",
}

# Standaard dunne rand voor de Excel-tabellen.
DUNNE_RANDEN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

alle_rijen = []        # Alle bruikbare rijen uit het bronbestand.
gefilterde_rijen = []  # Alleen lege, actieve A/B/C-locaties.
bron_bestand = None    # Het Excel-bestand dat nu is ingelezen.


def get_app_folder():
    # Bepaalt de map waar de applicatie draait, zowel als Python-script als als losse executable.
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def is_enabled(row):
    # Controleert of een locatie actief is volgens de kolom "Enabled?".
    return row.get("Enabled?") == 1


def is_valid_location(row):
    # Controleert of de locatie begint met een van de toegestane locatieletters.
    location = str(row.get("Location", "") or "")
    return location.startswith(GELDIGE_LOCATIONS)


def is_bron_excel_bestand(bestandsnaam):
    # Voorkomt dat eerder gegenereerde outputbestanden automatisch opnieuw als bronbestand worden ingelezen.
    return (
        bestandsnaam.endswith(".xlsx")
        and not bestandsnaam.startswith("Lege_locaties_")
        and not bestandsnaam.startswith("~$")
    )


def kolom_label(label):
    # Maakt WMS-kolomnamen vergelijkbaar, ook als er spaties of andere hoofdletters in zitten.
    return str(label or "").strip().casefold()


def vind_kolom(headers, canonieke_naam):
    # Zoekt de echte Excel-header die hoort bij de interne kolomnaam van de tool.
    mogelijke_namen = (canonieke_naam,) + KOLOM_ALIASSEN.get(canonieke_naam, ())
    for mogelijke_naam in mogelijke_namen:
        gezocht = kolom_label(mogelijke_naam)
        for header in headers:
            if kolom_label(header) == gezocht:
                return header
    return None


def maak_kolom_mapping(headers):
    # Koppelt actuele WMS-kolommen aan de vaste namen die de rest van het script gebruikt.
    mapping = {}
    ontbrekende_kolommen = []

    for kolom in VERPLICHTE_KOLOMMEN:
        gevonden_header = vind_kolom(headers, kolom)
        if gevonden_header is None:
            alias_tekst = ", ".join((kolom,) + KOLOM_ALIASSEN.get(kolom, ()))
            ontbrekende_kolommen.append(alias_tekst)
        else:
            mapping[kolom] = gevonden_header

    return mapping, ontbrekende_kolommen


def normaliseer_rij(headers, values, kolom_mapping):
    # Bewaart de originele kolommen en voegt de vaste interne kolomnamen toe.
    row = dict(zip(headers, values))
    for canonieke_naam, werkelijke_header in kolom_mapping.items():
        row[canonieke_naam] = row.get(werkelijke_header)
    return row


def laad_excel_bestand(file_path):
    # Leest een gekozen Excel-bestand in en werkt daarna de keuzelijsten en statusinformatie bij.
    global alle_rijen, bron_bestand

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(values_only=True), None)
            if header_row is None:
                alle_rijen = []
            else:
                headers = [str(h).strip() if h is not None else "" for h in header_row]
                kolom_mapping, ontbrekende_kolommen = maak_kolom_mapping(headers)
                if ontbrekende_kolommen:
                    raise ValueError(
                        "Dit lijkt niet het juiste bronbestand.\n"
                        f"Ontbrekende kolommen: {', '.join(ontbrekende_kolommen)}"
                    )
                alle_rijen = [
                    normaliseer_rij(headers, row, kolom_mapping)
                    for row in ws.iter_rows(min_row=2, values_only=True)
                ]
        finally:
            wb.close()

        bron_bestand = Path(file_path)
        process_data()
        export_button.config(state="normal")
        if bron_bestand.parent.resolve() == get_app_folder().resolve():
            status_var.set(f"\u2713 Geladen: {bron_bestand.name}")
            status_label.configure(foreground="#198754")
        else:
            status_var.set(f"Geladen: {bron_bestand.name}")
            status_label.configure(foreground="#000000")
    except Exception as e:
        messagebox.showerror("Fout", str(e))
        status_var.set("Kon het Excel-bestand niet laden.")
        status_label.configure(foreground="#B00020")


def laad_excel_automatisch():
    # Zoekt automatisch een bronbestand naast het script, zodat de gebruiker vaak meteen kan starten.
    script_map = get_app_folder()
    files = sorted(f for f in os.listdir(script_map) if is_bron_excel_bestand(f))
    if not files:
        status_var.set("Geen bronbestand gevonden. Kies handmatig een Excel-bestand.")
        return

    laad_excel_bestand(script_map / files[0])


def kies_excel_bestand():
    # Laat de gebruiker zelf een Excel-bestand kiezen wanneer het automatische bestand niet klopt.
    file_path = filedialog.askopenfilename(
        title="Kies Excel-bestand",
        initialdir=get_app_folder(),
        filetypes=[("Excel-bestanden", "*.xlsx"), ("Alle bestanden", "*.*")],
    )
    if file_path:
        laad_excel_bestand(file_path)


def process_data():
    # Zet ruwe Excel-rijen om naar bruikbare data met gang en locatienummer, en filtert op lege actieve locaties.
    global alle_rijen, gefilterde_rijen

    processed = []
    for row in alle_rijen:
        loc = str(row.get("Location", "") or "").strip()
        if not loc:
            continue

        gang = loc[:3]
        loc_num_str = loc[3:6]
        if not loc_num_str.isnumeric():
            continue

        processed_row = dict(row)
        processed_row.update(
            {
                "Location": loc,
                "Gang": gang,
                "LocatieNummer": int(loc_num_str),
            }
        )
        processed.append(processed_row)

    alle_rijen = processed
    gefilterde_rijen = [
        r
        for r in processed
        if is_valid_location(r)
        and r.get("Location Status") == EMPTY_STATUS
        and is_enabled(r)
    ]
    update_gangen()


def update_gangen():
    # Vult de gang-keuzelijst met alle unieke gangen uit de gefilterde lege locaties.
    unieke_gangen = sorted({r["Gang"] for r in gefilterde_rijen})
    gang_dropdown["values"] = [ALLE_GANGEN] + unieke_gangen
    gang_var.set(ALLE_GANGEN)


def get_location_type(loc):
    # Bepaalt welk uitzonderingsschema bij een locatie hoort op basis van de gangcode.
    gang = str(loc).strip()[:3]
    if gang.startswith("A"):
        return "A"
    if gang in ("BBX", "CCX"):
        return "BBX"
    if gang.startswith("B"):
        try:
            return "BC" if 10 <= int(gang[1:]) <= 21 else "OTHER"
        except ValueError:
            return "OTHER"
    if gang.startswith("C"):
        try:
            return "BC" if 22 <= int(gang[1:]) <= 32 else "OTHER"
        except ValueError:
            return "OTHER"
    return "OTHER"


def is_excluded_location(row):
    # Controleert of het locatienummer voorkomt in de uitzonderingslijst voor dit locatietype.
    location_type = get_location_type(row["Location"])
    return row["LocatieNummer"] in UITGEZONDERDE_LOCATIONS.get(location_type, set())


def houd_lege_locatie(row):
    # Houdt alleen lege locaties over die niet in de uitzonderingslijst staan.
    return not is_excluded_location(row)


def keep_full_juk_location(row):
    # Houdt volle locaties over die juist wel in de uitzonderingslijst staan.
    return is_excluded_location(row)


def get_family(sto_zone):
    # Herkent de familie van een locatie aan de hand van zoekwoorden in Sto_Zone_Cod.
    if not sto_zone:
        return None

    sto_zone_lower = str(sto_zone).lower()
    for keyword, label in zip(FAMILY_KEYWORDS, FAMILY_LABELS):
        if keyword in sto_zone_lower:
            return label
    return None


def get_height(row):
    # Leest de hoogte als getal; bij ontbrekende of ongeldige waarden wordt 0 gebruikt.
    try:
        return float(row.get("Height", 0) or 0)
    except (TypeError, ValueError):
        return 0


def build_family_rows():
    # Bouwt de brondata voor het family-overzicht met alleen lege, actieve locaties onder de maximale hoogte.
    family_rows = []
    for row in alle_rijen:
        if row.get("Location Status") != EMPTY_STATUS:
            continue
        if not is_enabled(row):
            continue

        aisle = str(row.get("Aisle", "") or "").strip()
        if not aisle:
            continue

        height = get_height(row)
        if height >= 1000:
            continue

        family = get_family(row.get("Sto_Zone_Cod", ""))
        if not family:
            continue

        family_rows.append({"Aisle": aisle, "Family": family, "Height": height})
    return family_rows


def apply_borders(ws):
    # Past de standaard dunne rand toe op alle cellen in een werkblad.
    for row in ws.iter_rows():
        for cell in row:
            cell.border = DUNNE_RANDEN


def append_family_totals(target_row, counts, key_prefix):
    # Voegt per familie het totale aantal toe aan een Excel-rij.
    for label in FAMILY_LABELS:
        family_count = counts[(*key_prefix, label)]
        target_row.append(family_count or "")


def create_empty_locations_sheet(ws, data_list):
    # Maakt het werkblad met de lijst van lege locaties en een kolom om af te vinken.
    ws.title = "lege locaties"
    ws.append(["Lege Locaties", "Hoogte", "Family type", "Afgevinkt"])
    for row in data_list:
        ws.append(
            [
                row["Location"],
                row.get("Height", ""),
                get_family(row.get("Sto_Zone_Cod", "")) or "",
                "",
            ]
        )
    apply_borders(ws)
    ws.auto_filter.ref = f"A1:D{ws.max_row}"


def create_family_availability_sheet(wb):
    # Maakt een overzicht per gang met totale locatiemogelijkheden en nog vrije locaties per family.
    ws = wb.create_sheet(title="family beschikbaar")
    ws.merge_cells("A1:H1")
    ws.merge_cells("J1:P1")
    ws.merge_cells("R1:X1")
    ws["A1"] = "TOTAAL AANTAL LOCATIES"
    ws["J1"] = "BESCHIKBAAR AANTAL LOCATIES"
    ws["R1"] = "BEZETTINGSGRAAD"
    ws.append(
        ["Aisle"]
        + list(FAMILY_LABELS)
        + ["TOTAAL", ""]
        + list(FAMILY_LABELS)
        + ["TOTAAL", ""]
        + list(FAMILY_LABELS)
        + ["TOTAAL"]
    )

    total_counts = Counter()
    available_counts = Counter()
    aisles = set()

    for row in alle_rijen:
        if row.get("Location Status") not in (EMPTY_STATUS, FULL_STATUS):
            continue
        if not is_valid_location(row) or not is_enabled(row):
            continue

        family = get_family(row.get("Sto_Zone_Cod", ""))
        if not family:
            continue

        aisle = row["Gang"]
        aisles.add(aisle)
        total_counts[(aisle, family)] += 1

    for row in gefilterde_rijen:
        if not houd_lege_locatie(row):
            continue

        family = get_family(row.get("Sto_Zone_Cod", ""))
        if not family:
            continue

        aisle = row["Gang"]
        aisles.add(aisle)
        available_counts[(aisle, family)] += 1

    total_row = ["TOTAAL"]
    total_values = [sum(total_counts[(aisle, family)] for aisle in aisles) for family in FAMILY_LABELS]
    available_values = [
        sum(available_counts[(aisle, family)] for aisle in aisles)
        for family in FAMILY_LABELS
    ]
    total_row += total_values
    total_row += [sum(total_values), ""]
    total_row += available_values + [sum(available_values), ""]
    total_row += [
        calculate_occupancy_rate(total_count, available_count)
        for total_count, available_count in zip(total_values, available_values)
    ]
    total_row += [calculate_occupancy_rate(sum(total_values), sum(available_values))]
    ws.append(total_row)

    for aisle in sorted(aisles):
        total_values = [total_counts[(aisle, family)] or "" for family in FAMILY_LABELS]
        available_values = [available_counts[(aisle, family)] or "" for family in FAMILY_LABELS]
        occupancy_values = [
            calculate_occupancy_rate(total_value, available_value)
            for total_value, available_value in zip(total_values, available_values)
        ]
        ws.append(
            [aisle]
            + total_values
            + [sum(value for value in total_values if value), ""]
            + available_values
            + [sum(value for value in available_values if value), ""]
            + occupancy_values
            + [
                calculate_occupancy_rate(
                    sum(value for value in total_values if value),
                    sum(value for value in available_values if value),
                )
            ]
        )

    style_family_availability_sheet(ws)


def calculate_occupancy_rate(total_count, available_count):
    # Berekent welk deel van de locaties bezet is; bij ontbrekende totalen blijft de cel leeg.
    if not total_count:
        return ""
    return (total_count - (available_count or 0)) / total_count


def create_full_locations_sheet(wb, data_full):
    # Maakt het werkblad met volle locaties die op een uitgezonderde derde locatie staan.
    ws = wb.create_sheet(title="pallet in 3e locatie")
    ws.append(["Gang", "Locatie", "Hoogte", "FULL"])
    for row in data_full:
        ws.append(
            [
                row["Gang"],
                row["LocatieNummer"],
                str(row["Location"])[-2:],
                row.get("Location Status", ""),
            ]
        )
    apply_borders(ws)


def create_family_sheet(wb, family_rows=None):
    # Maakt het family-overzicht met totaalaantallen per gang en familie.
    ws = wb.create_sheet(title="family overzicht")
    if family_rows is None:
        family_rows = build_family_rows()
    alle_aisles = sorted({r["Aisle"] for r in family_rows})
    counts = Counter()

    for row in family_rows:
        counts[(row["Aisle"], row["Family"])] += 1
        counts[(row["Family"],)] += 1

    header = [len(family_rows)] + list(FAMILY_LABELS)
    ws.append(header)

    style_family_header(ws)

    for aisle in alle_aisles:
        row_data = [aisle]
        append_family_totals(row_data, counts, (aisle,))
        ws.append(row_data)

    style_family_body(ws)

    total_row = ["TOTAAL"]
    append_family_totals(total_row, counts, ())
    ws.append(total_row)
    style_total_row(ws)
    configure_family_sheet_page(ws)


def format_height_label(height):
    # Toont hele hoogtes zonder decimalen en behoudt decimalen als die in de bron voorkomen.
    if float(height).is_integer():
        return str(int(height))
    return str(height)


def create_family_height_sheet(wb, family_rows=None):
    # Maakt een matrix met per gang de aantallen per family en exacte hoogte.
    ws = wb.create_sheet(title="family overzicht per hoogte")
    if family_rows is None:
        family_rows = build_family_rows()
    alle_aisles = sorted({r["Aisle"] for r in family_rows})
    heights_by_family = {
        family: sorted({r["Height"] for r in family_rows if r["Family"] == family})
        for family in FAMILY_LABELS
    }
    columns = [
        (family, height)
        for family in FAMILY_LABELS
        for height in heights_by_family[family]
    ]
    counts = Counter()

    for row in family_rows:
        counts[(row["Aisle"], row["Family"], row["Height"])] += 1
        counts[(row["Family"], row["Height"])] += 1

    ws.cell(row=1, column=1, value=len(family_rows))
    ws.cell(row=2, column=1, value="Gang")

    current_col = 2
    for family in FAMILY_LABELS:
        family_heights = heights_by_family[family]
        if not family_heights:
            continue

        start_col = current_col
        end_col = current_col + len(family_heights) - 1
        if start_col != end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col, value=family)

        for height in family_heights:
            ws.cell(row=2, column=current_col, value=format_height_label(height))
            current_col += 1

    for aisle in alle_aisles:
        row_data = [aisle]
        for family, height in columns:
            row_data.append(counts[(aisle, family, height)] or "")
        ws.append(row_data)

    total_row = ["TOTAAL"]
    for family, height in columns:
        total_row.append(counts[(family, height)] or "")
    ws.append(total_row)

    style_family_height_sheet(ws, columns)


def create_occupancy_sheet(wb):
    # Maakt het extra tabblad met totalen, vrije locaties, bezette locaties en bezettingsgraad per gang.
    ws = wb.create_sheet(title="bezettingsgraad")
    ws.append(["gang", "totaal aantal locaties", "bezet ", "vrij", "bezettingsgraad"])

    bezetting_per_gang = build_occupancy_rows()
    laatste_data_rij = len(bezetting_per_gang) + 2
    ws.append(
        [
            "",
            f"=SUM(B3:B{laatste_data_rij})",
            f"=SUM(C3:C{laatste_data_rij})",
            f"=SUM(D3:D{laatste_data_rij})",
            "=IFERROR(C2/B2,0)",
        ]
    )

    for gang, aantallen in bezetting_per_gang:
        totaal = aantallen[EMPTY_STATUS] + aantallen[FULL_STATUS]
        bezet = aantallen[FULL_STATUS]
        vrij = aantallen[EMPTY_STATUS]
        ws.append([gang, totaal, bezet, vrij, f"=IFERROR(C{ws.max_row + 1}/B{ws.max_row + 1},0)"])

    style_occupancy_sheet(ws, laatste_data_rij)


def build_occupancy_rows():
    # Telt per gang hoeveel ingeschakelde locaties vol en leeg zijn op basis van Enabled? en Location Status.
    counts = {}
    for row in alle_rijen:
        status = row.get("Location Status")
        if status not in (EMPTY_STATUS, FULL_STATUS):
            continue
        if not is_valid_location(row) or not is_enabled(row):
            continue

        gang = row["Gang"]
        if gang not in counts:
            counts[gang] = Counter()
        counts[gang][status] += 1

    return sorted(counts.items())


def style_occupancy_sheet(ws, laatste_data_rij):
    # Maakt het bezettingsgraad-tabblad op zoals het voorbeeldbestand.
    ws.column_dimensions["A"].width = 4.9
    ws.column_dimensions["B"].width = 18.8
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 10.3
    ws.column_dimensions["E"].width = 14.2
    ws.row_dimensions[2].height = 21

    for row in ws.iter_rows(min_row=1, max_row=laatste_data_rij, min_col=1, max_col=5):
        for cell in row:
            cell.border = DUNNE_RANDEN

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cell in ws[2][1:]:
        cell.font = Font(size=16)

    for row in ws.iter_rows(min_row=3, max_row=laatste_data_rij):
        row[0].font = Font(name="Arial", size=10, bold=True)
        row[0].alignment = Alignment(horizontal="center")
        row[4].number_format = "0.0%"

    ws["E2"].number_format = "0.0%"


def style_family_availability_sheet(ws):
    # Maakt het family-beschikbaar-tabblad scanbaar met een visuele scheiding tussen totaal en vrij.
    header_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    title_fill = PatternFill("solid", start_color="000000", end_color="000000")
    title_font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    aisle_fill = PatternFill("solid", start_color="ECEFF1", end_color="ECEFF1")
    total_fill = PatternFill("solid", start_color="CFD8DC", end_color="CFD8DC")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 4
    ws.column_dimensions["P"].width = 11
    ws.column_dimensions["Q"].width = 4
    ws.column_dimensions["X"].width = 11
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    for col_idx in range(2, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    for col_idx in range(10, 16):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    for col_idx in range(18, 24):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    for title_cell in (ws["A1"], ws["J1"], ws["R1"]):
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = DUNNE_RANDEN

    for cell in ws[2]:
        if cell.column in (9, 17):
            continue
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        cell.border = DUNNE_RANDEN

    for excel_row in ws.iter_rows(min_row=3):
        for cell in excel_row:
            if cell.column in (9, 17):
                continue

            cell.border = DUNNE_RANDEN
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)

            if cell.column == 1:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = aisle_fill
            elif 2 <= cell.column <= 7:
                color = FAMILY_KLEUREN[FAMILY_LABELS[cell.column - 2]]
                cell.fill = PatternFill("solid", start_color=color, end_color=color)
            elif cell.column in (8, 16, 24):
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = total_fill
            elif 10 <= cell.column <= 15:
                color = FAMILY_KLEUREN[FAMILY_LABELS[cell.column - 10]]
                cell.fill = PatternFill("solid", start_color=color, end_color=color)
            elif 18 <= cell.column <= 23:
                color = FAMILY_KLEUREN[FAMILY_LABELS[cell.column - 18]]
                cell.fill = PatternFill("solid", start_color=color, end_color=color)

            if 18 <= cell.column <= 24:
                cell.number_format = "0.0%"

            if cell.row == 3:
                cell.font = Font(name="Arial", size=10, bold=True)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.2,
        bottom=0.2,
        header=0.1,
        footer=0.1,
    )


def style_family_header(ws):
    # Geeft de kopregel van het family-overzicht een donkere achtergrond, witte tekst en randen.
    header_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = DUNNE_RANDEN


def style_family_body(ws):
    # Maakt de datarijen van het family-overzicht op en kleurt de familiekolommen.
    for excel_row in ws.iter_rows(min_row=2):
        for col_idx, cell in enumerate(excel_row):
            cell.border = DUNNE_RANDEN
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)

            if col_idx == 0:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = PatternFill("solid", start_color="ECEFF1", end_color="ECEFF1")
                continue

            family_idx = col_idx - 1
            if 0 <= family_idx < len(FAMILY_LABELS):
                color = FAMILY_KLEUREN[FAMILY_LABELS[family_idx]]
                cell.fill = PatternFill("solid", start_color=color, end_color=color)


def style_total_row(ws):
    # Geeft de totaalregel onderaan het family-overzicht dezelfde opvallende stijl als de kopregel.
    total_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    total_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = DUNNE_RANDEN


def configure_family_sheet_page(ws):
    # Stelt kolombreedtes en pagina-instellingen in zodat het family-overzicht netjes op een pagina past.
    ws.column_dimensions["A"].width = 10
    ws.row_dimensions[1].height = 30

    for i in range(len(FAMILY_LABELS)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 11

    ws.print_area = f"A1:G{ws.max_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 155
    ws.page_setup.fitToWidth = False
    ws.page_setup.fitToHeight = False
    ws.page_margins = PageMargins(
        left=0.1,
        right=0.1,
        top=0.1,
        bottom=0.1,
        header=0.1,
        footer=0.1,
    )
    ws.print_options.horizontalCentered = True


def style_family_height_sheet(ws, columns):
    # Maakt het hoogte-overzicht op met dezelfde family-kleuren als het compacte overzicht.
    header_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    aisle_fill = PatternFill("solid", start_color="ECEFF1", end_color="ECEFF1")
    total_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    total_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 10
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            if cell.value in (None, ""):
                continue
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = DUNNE_RANDEN

    for excel_row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in excel_row:
            cell.border = DUNNE_RANDEN
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)

            if cell.column == 1:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = aisle_fill
                continue

            family = columns[cell.column - 2][0]
            color = FAMILY_KLEUREN[family]
            cell.fill = PatternFill("solid", start_color=color, end_color=color)

    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = DUNNE_RANDEN

    for family_idx, family in enumerate(family for family in FAMILY_LABELS if any(col[0] == family for col in columns)):
        if family_idx == 0:
            continue
        start_col = next(idx + 2 for idx, col in enumerate(columns) if col[0] == family)
        ws.column_dimensions[get_column_letter(start_col)].width = 10

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.2,
        bottom=0.2,
        header=0.1,
        footer=0.1,
    )


def get_empty_locations_for_selection():
    # Past de huidige gang- en even/oneven-keuze toe en geeft de exporteerbare lege locaties terug.
    gang_selectie = gang_var.get()
    keuze = keuze_var.get()

    if gang_selectie == ALLE_GANGEN:
        data_list = [r for r in gefilterde_rijen if houd_lege_locatie(r)]
    else:
        data_list = [
            r
            for r in gefilterde_rijen
            if r["Gang"] == gang_selectie and houd_lege_locatie(r)
        ]

    if keuze == "Even":
        data_list = [r for r in data_list if r["LocatieNummer"] % 2 == 0]
    elif keuze == "Oneven":
        data_list = [r for r in data_list if r["LocatieNummer"] % 2 == 1]

    data_list.sort(key=lambda r: (r["Gang"], r["LocatieNummer"]))
    return data_list


def get_full_juk_locations():
    # Verzamelt volle locaties die in de speciale uitzonderingsposities vallen.
    data_full = [
        r
        for r in alle_rijen
        if r.get("Location Status") == FULL_STATUS
        and is_valid_location(r)
        and is_enabled(r)
        and keep_full_juk_location(r)
    ]
    data_full.sort(key=lambda r: (r["Gang"], r["LocatieNummer"]))
    return data_full


def maak_uniek_output_pad(output_path):
    # Maakt een unieke bestandsnaam als er al een exportbestand met dezelfde naam bestaat.
    if not output_path.exists():
        return output_path

    for nummer in range(2, 100):
        kandidaat = output_path.with_name(f"{output_path.stem}_{nummer}{output_path.suffix}")
        if not kandidaat.exists():
            return kandidaat

    return output_path


def filter_and_export():
    # Filtert de ingelezen data op gang en even/oneven keuze, maakt de Excel-werkmap en slaat die op.
    if not gefilterde_rijen:
        messagebox.showinfo("Geen data", "Er zijn nog geen lege locaties geladen om te exporteren.")
        return

    gang_selectie = gang_var.get()
    keuze = keuze_var.get()
    data_list = get_empty_locations_for_selection()
    if not data_list:
        messagebox.showinfo("Geen resultaat", "Er zijn geen lege locaties voor deze selectie.")
        return

    data_full = get_full_juk_locations()

    try:
        g_naam = "Alle_Gangen" if gang_selectie == ALLE_GANGEN else gang_selectie
        bestandsnaam = f"Lege_locaties_{g_naam}_{keuze}.xlsx"
        output_path = maak_uniek_output_pad(get_app_folder() / bestandsnaam)
        family_rows = build_family_rows()

        wb = Workbook()
        create_empty_locations_sheet(wb.active, data_list)
        create_family_sheet(wb, family_rows)
        create_family_height_sheet(wb, family_rows)
        create_family_availability_sheet(wb)
        create_occupancy_sheet(wb)
        create_full_locations_sheet(wb, data_full)

        wb.save(output_path)
        os.startfile(output_path)
        status_var.set(f"Bestand opgeslagen en geopend: {output_path.name}")
    except Exception as e:
        messagebox.showerror("Export fout", str(e))


# Bouwt het hoofdvenster van de Tkinter-applicatie.
root = tk.Tk()
root.title("Locatie Tool")
root.geometry("620x320")
root.resizable(False, False)

# Plaatst het hoofdframe met extra ruimte voor labels, status en filterinformatie.
frame = ttk.Frame(root, padding=20)
frame.pack(fill="both", expand=True)

# Maakt Tkinter-variabelen voor de geselecteerde gang, even/oneven keuze en statusregels.
gang_var = tk.StringVar()
keuze_var = tk.StringVar(value="Beide")
status_var = tk.StringVar(value="Bestand laden...")

# Toont een titel en korte context boven de bedieningselementen.
ttk.Label(frame, text="Lege locaties exporteren", font=("Arial", 14, "bold")).grid(
    row=0,
    column=0,
    columnspan=4,
    sticky="w",
    pady=(0, 12),
)

# Maakt een knop waarmee de gebruiker zelf het bronbestand kan kiezen.
ttk.Button(frame, text="Kies Excel-bestand", command=kies_excel_bestand).grid(
    row=1,
    column=0,
    sticky="w",
    pady=(0, 12),
)

# Maakt de keuzelijst voor de gang; de waarden worden gevuld nadat het Excel-bestand is ingelezen.
ttk.Label(frame, text="Gang").grid(row=2, column=0, sticky="w")
gang_dropdown = ttk.Combobox(frame, textvariable=gang_var, width=15, state="readonly")
gang_dropdown.grid(row=3, column=0, sticky="w", padx=(0, 12), pady=(2, 12))

# Maakt de keuzelijst waarmee de gebruiker even, oneven of beide locatienummers kan kiezen.
ttk.Label(frame, text="Locatienummers").grid(row=2, column=1, sticky="w")
keuze_dropdown = ttk.Combobox(
    frame,
    textvariable=keuze_var,
    values=["Even", "Oneven", "Beide"],
    width=10,
    state="readonly",
)
keuze_dropdown.grid(row=3, column=1, sticky="w", padx=(0, 12), pady=(2, 12))

# Maakt de knop die de gefilterde data exporteert naar een nieuw Excel-bestand.
export_button = ttk.Button(frame, text="Genereer Excel", command=filter_and_export, state="disabled")
export_button.grid(
    row=3,
    column=2,
    sticky="w",
    padx=(0, 12),
    pady=(2, 12),
)

# Toont de laadstatus van het Excel-bestand onder de bedieningselementen.
status_label = ttk.Label(frame, textvariable=status_var)
status_label.grid(row=4, column=0, columnspan=4, sticky="w", pady=(8, 2))

# Start kort na het openen automatisch het inlezen van Excel en daarna de Tkinter-eventloop.
root.after(50, laad_excel_automatisch)
root.mainloop()
