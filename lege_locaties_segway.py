import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side

rows_raw = []       # Alle rijen als lijst van dicts
rows_filtered = []  # Gefilterde rijen (A/B/C + Empty)

# =========================
# Excel automatisch laden
# =========================
def load_excel_auto():
    global rows_raw

    if getattr(sys, 'frozen', False):
        script_map = os.path.dirname(sys.executable)
    else:
        script_map = os.path.dirname(os.path.abspath(__file__))

    try:
        files = [f for f in os.listdir(script_map) if f.endswith(".xlsx")]

        if not files:
            messagebox.showerror("Fout", "Geen Excel bestand gevonden!")
            return

        file_path = os.path.join(script_map, files[0])

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        headers = []
        rows_raw = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h) if h is not None else "" for h in row]
                if 'Location' not in headers or 'Location Status' not in headers:
                    messagebox.showerror("Fout", "Je hebt je vorige bestand nog niet verwijderd!")
                    wb.close()
                    return
            else:
                row_dict = dict(zip(headers, row))
                rows_raw.append(row_dict)

        wb.close()
        process_data()
        status_label.config(text=f"Loaded: {files[0]}")

    except Exception as e:
        messagebox.showerror("Fout", str(e))


# =========================
# Data verwerken
# =========================
def process_data():
    global rows_raw, rows_filtered

    processed = []
    for row in rows_raw:
        loc = str(row.get('Location', '') or '').strip()
        if not loc:
            continue

        gang = loc[:3]
        loc_num_str = loc[3:6]

        if not loc_num_str.isnumeric():
            continue

        loc_num = int(loc_num_str)
        row = dict(row)  # kopie
        row['Location'] = loc
        row['Gang'] = gang
        row['LocatieNummer'] = loc_num
        processed.append(row)

    rows_raw = processed

    rows_filtered = [
        r for r in processed
        if r['Location'][0] in ('A', 'B', 'C')
        and r.get('Location Status') == "E - Empty"
    ]

    update_gangen()


# =========================
# Gangen vullen
# =========================
def update_gangen():
    gangen = sorted(set(r['Gang'] for r in rows_filtered))
    gang_dropdown['values'] = gangen

    if gangen:
        gang_var.set(gangen[0])


# =========================
# TYPE DETECTIE
# Gang (eerste 3 tekens) bepaalt het type:
#   A-gangen          → "A"      → exclude_locations_A
#   B10 t/m B21       → "BC"     → exclude_locations_BC
#   C22 t/m C32       → "BC"     → exclude_locations_BC
#   BBX               → "BBX"    → exclude_locations_BBX
#   CCX               → "BBX"    → exclude_locations_BBX
#   overig            → "OTHER"
# =========================
def get_location_type(loc):
    loc = str(loc).strip()
    gang = loc[:3]

    if gang.startswith("A"):
        return "A"
    elif gang in ("BBX",):
        return "BBX"
    elif gang in ("CCX",):
        return "BBX"
    elif gang.startswith("B"):
        # B10 t/m B21
        try:
            gang_num = int(gang[1:])
            if 10 <= gang_num <= 21:
                return "BC"
        except ValueError:
            pass
        return "OTHER"
    elif gang.startswith("C"):
        # C22 t/m C32
        try:
            gang_num = int(gang[1:])
            if 22 <= gang_num <= 32:
                return "BC"
        except ValueError:
            pass
        return "OTHER"
    else:
        return "OTHER"


# =========================
# Filter + EXPORT
# =========================
def filter_and_export():
    if not rows_filtered:
        messagebox.showwarning("Let op", "Geen data geladen")
        return

    gang = gang_var.get()
    keuze = keuze_var.get()

    # =========================
    # EXCLUDE LISTS
    # =========================
    exclude_locations_BC = {
        5, 6, 13, 14, 21, 22, 29, 30, 37, 38, 45, 46, 53, 54, 61, 62,
        69, 70, 77, 78, 85, 86, 93, 94, 101, 102, 109, 110, 117, 118,
        125, 126, 133, 134, 141, 142, 149, 150, 157, 158, 165, 166,
        173, 174, 181, 182, 189, 190, 197, 198
    }

    exclude_locations_A = {
        9, 10, 17, 18, 25, 26, 33, 34, 41, 42, 49, 50,
        57, 58, 65, 66, 73, 74, 81, 82, 89, 90, 97, 98,
        105, 106, 113, 114, 121, 122, 129, 130, 137, 138,
        145, 146, 153, 154, 161, 162, 169, 170, 177, 178,
        185, 186, 193, 194, 201, 202, 209, 210
    }

    exclude_locations_BBX = {3, 7, 11, 15, 19, 23, 27, 31, 35, 39, 43, 47, 51, 55, 59, 63, 67, 71, 75, 79}

    # =========================
    # EXCLUDE LOGICA
    # =========================
    def apply_exclude(row):
        loc_type = get_location_type(row['Location'])
        loc_num = row['LocatieNummer']

        if loc_type == "A":
            return loc_num not in exclude_locations_A
        elif loc_type == "BBX":
            return loc_num not in exclude_locations_BBX
        elif loc_type == "BC":
            return loc_num not in exclude_locations_BC
        else:
            return False


    # =========================
    # DATA VOOR LIJST (Sheet 1)
    # =========================
    data_list = [
        r for r in rows_filtered
        if r['Gang'] == gang
        and apply_exclude(r)
    ]

    if keuze == "Even":
        data_list = [r for r in data_list if r['LocatieNummer'] % 2 == 0]
    elif keuze == "Oneven":
        data_list = [r for r in data_list if r['LocatieNummer'] % 2 == 1]

    data_list.sort(key=lambda r: r['LocatieNummer'])

    if not data_list:
        messagebox.showwarning("Leeg", "Geen resultaten")
        return

    # =========================
    # DATA VOOR OVERZICHT (Sheet 2)
    # =========================
    data_all = [
        r for r in rows_filtered
        if apply_exclude(r)
    ]

    if keuze == "Even":
        data_all = [r for r in data_all if r['LocatieNummer'] % 2 == 0]
    elif keuze == "Oneven":
        data_all = [r for r in data_all if r['LocatieNummer'] % 2 == 1]

    # =========================
    # DATA VOOR FULL JUK 3 (Sheet 3)
    # =========================
    def apply_full_filter(row):
        loc_type = get_location_type(row['Location'])
        loc_num = row['LocatieNummer']

        if loc_type == "A":
            return loc_num in exclude_locations_A
        elif loc_type == "BBX":
            return loc_num in exclude_locations_BBX
        elif loc_type == "BC":
            return loc_num in exclude_locations_BC
        else:
            return False

    data_full = [
        r for r in rows_raw
        if r.get('Location Status') == "F - Full"
        and r['Location'][0] in ('A', 'B', 'C')
        and apply_full_filter(r)
    ]
    data_full.sort(key=lambda r: (r['Gang'], r['LocatieNummer']))

    # =========================
    # EXPORT
    # =========================
    try:
        if getattr(sys, 'frozen', False):
            base_path = os.path.dirname(sys.executable)
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))

        bestandsnaam = f"Lege_locaties_{gang}_{keuze}.xlsx"
        output_path = os.path.join(base_path, bestandsnaam)

        wb = Workbook()

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def apply_borders(ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        # ---------------------------
        # SHEET 1 – Lijst
        # ---------------------------
        ws1 = wb.active
        ws1.title = "Lijst"
        ws1.append(["Lege Locaties", "Afgevinkt"])

        for row in data_list:
            ws1.append([row['Location'], ""])

        apply_borders(ws1)

        ws1.page_setup.fitToWidth = 1
        ws1.page_setup.fitToHeight = False
        ws1.page_margins.top = 0
        ws1.page_margins.bottom = 0
        ws1.page_margins.left = 1
        ws1.page_margins.right = 0

        # ---------------------------
        # SHEET 2 – Overzicht (pivot)
        # ---------------------------
        ws2 = wb.create_sheet(title="Overzicht")

        # Controleer of Height-kolom beschikbaar is
        has_height = any('Height' in r for r in data_all)

        if has_height:
            # Handmatige pivot: rijen = Gang, kolommen = Height, waarden = count(Location)
            pivot = {}      # {gang: {height: count}}
            heights = set()

            for r in data_all:
                g = r['Gang']
                h = r.get('Height', '')
                pivot.setdefault(g, {})
                pivot[g][h] = pivot[g].get(h, 0) + 1
                heights.add(h)

            sorted_heights = sorted(heights, key=lambda x: (x is None, x))
            header = ['Gang'] + sorted_heights
            ws2.append(header)

            for g in sorted(pivot.keys()):
                row_out = [g] + [pivot[g].get(h, 0) for h in sorted_heights]
                ws2.append(row_out)

        apply_borders(ws2)

        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = False
        ws2.page_margins.top = 0
        ws2.page_margins.bottom = 0
        ws2.page_margins.left = 0
        ws2.page_margins.right = 0

        # ---------------------------
        # SHEET 3 – Full Juk 3
        # ---------------------------
        ws3 = wb.create_sheet(title="Full_Juk_3")
        ws3.append(["Gang", "Locatie", "Hoogte", "FULL"])

        for r in data_full:
            location = str(r['Location'])
            hoogte = location[-2:] if location[-2:].isdigit() else ""
            ws3.append([
                r['Gang'],
                r['LocatieNummer'],
                hoogte,
                r.get('Location Status', '')
            ])

        apply_borders(ws3)

        ws3.page_setup.fitToWidth = 1
        ws3.page_setup.fitToHeight = False
        ws3.page_margins.top = 0
        ws3.page_margins.bottom = 0
        ws3.page_margins.left = 1
        ws3.page_margins.right = 0

        wb.save(output_path)
        messagebox.showinfo("Succes", f"Bestand opgeslagen:\n{bestandsnaam}")

    except Exception as e:
        messagebox.showerror("Export fout", str(e))


# =========================
# GUI
# =========================
root = tk.Tk()
root.title("Locatie Tool")
root.geometry("500x250")

frame = tk.Frame(root)
frame.pack(pady=20)

gang_var = tk.StringVar()
gang_dropdown = ttk.Combobox(frame, textvariable=gang_var, width=10)
gang_dropdown.grid(row=0, column=0, padx=5)

keuze_var = tk.StringVar(value="Beide")
ttk.Combobox(
    frame,
    textvariable=keuze_var,
    values=["Even", "Oneven", "Beide"],
    width=10
).grid(row=0, column=1, padx=5)

tk.Button(frame, text="Genereer Excel", command=filter_and_export).grid(row=0, column=2, padx=5)

status_label = tk.Label(root, text="Bestand laden...")
status_label.pack(pady=10)

root.after(100, load_excel_auto)
root.mainloop()
