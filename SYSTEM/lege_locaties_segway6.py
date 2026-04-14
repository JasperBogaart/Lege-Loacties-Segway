import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side

rows_raw = []       # Alle rijen als lijst van dicts
rows_filtered = []  # Gefilterde rijen (A/B/C + Empty)

# De lijsten met locatienummers
buitenste_locaties_B_C = [1, 2, 7, 8, 9, 10, 15, 16, 17, 18, 23, 24, 25, 26, 31, 32, 33, 34,
    39, 40, 41, 42, 47, 48, 49, 50, 55, 56, 57, 58, 63, 64, 65, 66, 71, 72, 73, 74,
    79, 80, 81, 82, 87, 88, 89, 90, 95, 96, 97, 98, 103, 104, 105, 106, 111, 112, 113, 114,
    119, 120, 121, 122, 127, 128, 129, 130, 135, 136, 137, 138, 143, 144, 145, 146, 151, 152, 153, 154,
    159, 160, 161, 162, 167, 168, 169, 170, 175, 176, 177, 178, 183, 184, 185, 186, 191, 192, 193, 194, 199, 200
]

middelste_loacties_B_C = [3, 4, 11, 12, 19, 20, 27, 28, 35, 36, 43, 44, 51, 52, 59, 60, 67, 68,
    75, 76, 83, 84, 91, 92, 99, 100, 107, 108, 115, 116, 123, 124, 131, 132, 139, 140, 147, 148,
    155, 156, 163, 164, 171, 172, 179, 180, 187, 188, 195, 196
]

buitenste_locaties_A = [1, 2, 3, 4, 5, 6, 11, 12, 13, 14, 19, 20, 21, 22, 27, 28, 29, 30, 35, 36, 37, 38, 43, 44,
    45, 46, 51, 52, 53, 54, 59, 60, 61, 62, 67, 68, 69, 70, 75, 76, 77, 78, 83, 84, 85, 86, 91, 92, 93, 94, 99, 100,
    101, 102, 107, 108, 109, 110, 115, 116, 117, 118, 123, 124, 125, 126, 131, 132, 133, 134, 139, 140, 141, 142, 147,
    148, 149, 150, 155, 156, 157, 158, 163, 164, 165, 166, 171, 172, 173, 174, 179, 180, 181, 182, 187, 188, 189, 190,
    195, 196, 197, 198, 203, 204, 205, 206, 211, 212
]

middelste_loacties_A = [7, 8, 15, 16, 23, 24, 31, 32, 39, 40, 47, 48, 55, 56, 63, 64, 71, 72, 79, 80, 87, 88, 95, 96,
    103, 104, 111, 112, 119, 120, 127, 128, 135, 136, 143, 144, 151, 152, 159, 160, 167, 168, 175, 176, 183, 184, 191,
    192, 199, 200, 207, 208
]

buitenste_locaties_CCX = [1, 4, 5, 8, 9, 12, 13, 16, 17, 20, 21, 24, 25, 28, 29, 32, 33, 36, 37, 40, 41, 44, 45, 48,
    49, 52, 53, 56, 57, 60, 61, 64
]

middelste_loacties_CCX = [2, 6, 10, 14, 18, 22, 26, 30, 34, 38, 42, 46, 50, 54, 58, 62]

def get_position_sets(gang):
    gang = str(gang).strip()
    loc_type = get_location_type(gang + "001")
    if loc_type == "A": return set(buitenste_locaties_A), set(middelste_loacties_A)
    elif loc_type == "BC": return set(buitenste_locaties_B_C), set(middelste_loacties_B_C)
    elif loc_type == "BBX":
        if gang == "CCX": return set(buitenste_locaties_CCX), set(middelste_loacties_CCX)
    return set(), set()

def load_excel_auto():
    global rows_raw
    if getattr(sys, 'frozen', False): script_map = os.path.dirname(sys.executable)
    else: script_map = os.path.dirname(os.path.abspath(__file__))
    try:
        files = [f for f in os.listdir(script_map) if f.endswith(".xlsx")]
        if not files: return
        file_path = os.path.join(script_map, files[0])
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(h) if h is not None else "" for h in next(ws.iter_rows(values_only=True))]
        rows_raw = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
        wb.close()
        process_data()
        status_label.config(text=f"Loaded: {files[0]}")
    except Exception as e: messagebox.showerror("Fout", str(e))

def process_data():
    global rows_raw, rows_filtered
    processed = []
    for row in rows_raw:
        loc = str(row.get('Location', '') or '').strip()
        if not loc: continue
        gang, loc_num_str = loc[:3], loc[3:6]
        if not loc_num_str.isnumeric(): continue
        row = dict(row)
        row.update({'Location': loc, 'Gang': gang, 'LocatieNummer': int(loc_num_str)})
        processed.append(row)
    rows_raw = processed
    rows_filtered = [r for r in processed if r['Location'][0] in ('A', 'B', 'C') and r.get('Location Status') == "E - Empty" and r.get('Enabled?') == 1]
    update_gangen()

def update_gangen():
    # Haal alle unieke gangen op uit de gefilterde data
    unieke_gangen = sorted(list(set(r['Gang'] for r in rows_filtered)))
    # Maak de nieuwe lijst met "Alle Gangen" als eerste optie
    dropdown_opties = ["Alle Gangen"] + unieke_gangen
    # Update de dropdown waarden
    gang_dropdown['values'] = dropdown_opties
    # Zet de standaard geselecteerde waarde
    gang_var.set("Alle Gangen")

def get_location_type(loc):
    loc, gang = str(loc).strip(), str(loc)[:3]
    if gang.startswith("A"): return "A"
    elif gang in ("BBX", "CCX"): return "BBX"
    elif gang.startswith("B"):
        try:
            if 10 <= int(gang[1:]) <= 21: return "BC"
        except: pass
    elif gang.startswith("C"):
        try:
            if 22 <= int(gang[1:]) <= 32: return "BC"
        except: pass
    return "OTHER"

def filter_and_export():
    if not rows_filtered: return
    gang_selectie, keuze = gang_var.get(), keuze_var.get()

    exclude_locations_BC = {5, 6, 13, 14, 21, 22, 29, 30, 37, 38, 45, 46, 53, 54, 61, 62, 69, 70, 77, 78, 85, 86, 93, 94, 101, 102, 109, 110, 117, 118, 125, 126, 133, 134, 141, 142, 149, 150, 157, 158, 165, 166, 173, 174, 181, 182, 189, 190, 197, 198}
    exclude_locations_A = {9, 10, 17, 18, 25, 26, 33, 34, 41, 42, 49, 50, 57, 58, 65, 66, 73, 74, 81, 82, 89, 90, 97, 98, 105, 106, 113, 114, 121, 122, 129, 130, 137, 138, 145, 146, 153, 154, 161, 162, 169, 170, 177, 178, 185, 186, 193, 194, 201, 202, 209, 210}
    exclude_locations_BBX = {3, 7, 11, 15, 19, 23, 27, 31, 35, 39, 43, 47, 51, 55, 59, 63, 67, 71, 75, 79}

    def apply_exclude(row):
        lt, ln = get_location_type(row['Location']), row['LocatieNummer']
        if lt == "A": return ln not in exclude_locations_A
        if lt == "BBX": return ln not in exclude_locations_BBX
        if lt == "BC": return ln not in exclude_locations_BC
        return False

    # Filteren voor de "Lijst" sheet
    if gang_selectie == "Alle Gangen":
        data_list = [r for r in rows_filtered if apply_exclude(r)]
    else:
        data_list = [r for r in rows_filtered if r['Gang'] == gang_selectie and apply_exclude(r)]

    if keuze == "Even": data_list = [r for r in data_list if r['LocatieNummer'] % 2 == 0]
    elif keuze == "Oneven": data_list = [r for r in data_list if r['LocatieNummer'] % 2 == 1]
    
    data_list.sort(key=lambda r: (r['Gang'], r['LocatieNummer']))

    data_all = [r for r in rows_filtered if apply_exclude(r)]
    if keuze == "Even": data_all = [r for r in data_all if r['LocatieNummer'] % 2 == 0]
    elif keuze == "Oneven": data_all = [r for r in data_all if r['LocatieNummer'] % 2 == 1]

    def is_full_juk_target(row):
        lt, ln = get_location_type(row['Location']), row['LocatieNummer']
        if lt == "A": return ln in exclude_locations_A
        if lt == "BBX": return ln in exclude_locations_BBX
        if lt == "BC": return ln in exclude_locations_BC
        return False

    data_full = [r for r in rows_raw if r.get('Location Status') == "F - Full" and r['Location'][0] in ('A', 'B', 'C') and r.get('Enabled?') == 1 and is_full_juk_target(r)]
    data_full.sort(key=lambda r: (r['Gang'], r['LocatieNummer']))

    alle_gangen = sorted(set(r['Gang'] for r in rows_filtered))
    lege_overzicht = []
    for g in alle_gangen:
        bs, ms = get_position_sets(g)
        if not bs and not ms: continue
        b_laag, b_hoog, m_laag, m_hoog = 0, 0, 0, 0
        for r in [r for r in rows_filtered if r['Gang'] == g]:
            try: h = float(r.get('Height', 0) or 0)
            except: h = 0
            if r['LocatieNummer'] in bs:
                if h < 237: b_laag += 1
                else: b_hoog += 1
            elif r['LocatieNummer'] in ms:
                if h < 237: m_laag += 1
                else: m_hoog += 1
        lege_overzicht.append((g, b_laag, b_hoog, m_laag, m_hoog))

    try:
        g_naam = "Alle_Gangen" if gang_selectie == "Alle Gangen" else gang_selectie
        bestandsnaam = f"Lege_locaties_{g_naam}_{keuze}.xlsx"
        output_path = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__)), bestandsnaam)
        
        wb = Workbook()
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def apply_borders(ws):
            for row in ws.iter_rows():
                for cell in row: cell.border = thin_border

        ws1, ws2, ws3 = wb.active, wb.create_sheet(title="Overzicht"), wb.create_sheet(title="Full_Juk_3")
        ws1.title = "Lijst"
        ws1.append(["Lege Locaties", "Afgevinkt"])
        for r in data_list: ws1.append([r['Location'], ""])
        apply_borders(ws1)
        
        heights = sorted({r.get('Height', '') for r in data_all}, key=lambda x: (x is None, x))
        ws2.append(['Gang'] + heights)
        pivot = {}
        for r in data_all:
            g, h = r['Gang'], r.get('Height', '')
            pivot.setdefault(g, {})[h] = pivot.setdefault(g, {}).get(h, 0) + 1
        for g in sorted(pivot.keys()): ws2.append([g] + [pivot[g].get(h, 0) for h in heights])
        apply_borders(ws2)

        ws3.append(["Gang", "Locatie", "Hoogte", "FULL"])
        for r in data_full: ws3.append([r['Gang'], r['LocatieNummer'], str(r['Location'])[-2:], r.get('Location Status', '')])
        apply_borders(ws3)

        ws4 = wb.create_sheet(title="Lege_Locaties")
        ws4.append(["Gang", "Buitenste (< 237)", "Buitenste (> 237)", "Middelste (< 237)", "Middelste (> 237)"])
        for row in lege_overzicht: ws4.append(row)
        apply_borders(ws4)
        for col in ['B', 'C', 'D', 'E']: ws4.column_dimensions[col].width = 14

        wb.save(output_path)
        messagebox.showinfo("Succes", f"Bestand opgeslagen:\n{bestandsnaam}")
    except Exception as e: messagebox.showerror("Export fout", str(e))

# GUI
root = tk.Tk()
root.title("Locatie Tool")
root.geometry("500x250")
frame = tk.Frame(root)
frame.pack(pady=20)
gang_var, keuze_var = tk.StringVar(), tk.StringVar(value="Beide")
# Breedte aangepast naar 15 zodat "Alle Gangen" goed zichtbaar is
gang_dropdown = ttk.Combobox(frame, textvariable=gang_var, width=15, state="readonly")
gang_dropdown.grid(row=0, column=0, padx=5)
ttk.Combobox(frame, textvariable=keuze_var, values=["Even", "Oneven", "Beide"], width=10, state="readonly").grid(row=0, column=1, padx=5)
tk.Button(frame, text="Genereer Excel", command=filter_and_export).grid(row=0, column=2, padx=5)
status_label = tk.Label(root, text="Bestand laden...")
status_label.pack(pady=10)
root.after(100, load_excel_auto)
root.mainloop()