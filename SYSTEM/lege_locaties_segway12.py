import os
import sys
import tkinter as tk
from collections import Counter
from pathlib import Path
from tkinter import filedialog, messagebox, ttk


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


# Definieert vaste teksten en geldige locatiegroepen die door de hele tool worden gebruikt.
ALLE_GANGEN = "Alle Gangen"
EMPTY_STATUS = "E - Empty"
FULL_STATUS = "F - Full"
GELDIGE_LOCATIONS = ("A", "B", "C")
VERPLICHTE_KOLOMMEN = ("Location", "Location Status", "Enabled?", "Aisle", "Height", "Sto_Zone_Cod")

# Legt vast welke locatienummers worden uitgezonderd bij lege locaties of juist apart worden getoond bij volle locaties.
UITGEZONDERDE_LOCATIONS = {
    "A": set(range(9, 211, 8)) | set(range(10, 211, 8)),
    "BBX": set(range(3, 80, 4)),
    "BC": set(range(5, 199, 8)) | set(range(6, 199, 8)),
}

# Koppelt zoekwoorden uit Sto_Zone_Cod aan leesbare familienamen en kleuren voor het family-overzicht.
FAMILY_KEYWORDS = ("big", "med-lng", "med-shrt", "sml-lng", "sml-shrt", "ground")
FAMILY_LABELS = ("BIG", "MED-LNG", "MED-SHRT", "SML-LNG", "SML-SHRT", "GROUND")
FAMILY_KLEUREN = {
    "BIG": "D9EAD3",
    "MED-LNG": "CFE2F3",
    "MED-SHRT": "EAD1DC",
    "SML-LNG": "FFF2CC",
    "SML-SHRT": "F4CCCC",
    "GROUND": "E6E0EC",
}

# Maakt een herbruikbare dunne celrand voor alle Excel-tabellen.
DUNNE_RANDEN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

alle_rijen = []       # Alle rijen als lijst van dicts
gefilterde_rijen = []  # Gefilterde rijen (A/B/C + Empty)
bron_bestand = None   # Het Excel-bestand dat op dit moment is ingelezen


def get_app_folder():
    # Bepaalt de map waar de applicatie draait, zowel als Python-script als als losse executable.
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def is_enabled(row):
    # Controleert of een locatie actief is volgens de kolom "Enabled?".
    return row.get("Enabled?") == 1


def is_valid_location(row):
    # Controleert of de locatie begint met een van de toegestane locatieletters.
    location = str(row.get("Location", "") or "")
    return location.startswith(GELDIGE_LOCATIONS)


def is_bron_excel_bestand(bestandsnaam):
    # Voorkomt dat eerder gegenereerde outputbestanden automatisch opnieuw als bronbestand worden ingelezen.
    return (
        bestandsnaam.endswith(".xlsx")
        and not bestandsnaam.startswith("Lege_locaties_")
        and not bestandsnaam.startswith("~$")
    )


def laad_excel_bestand(file_path):
    # Leest een gekozen Excel-bestand in en werkt daarna de keuzelijsten en statusinformatie bij.
    global alle_rijen, bron_bestand

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(values_only=True), None)
            if header_row is None:
                alle_rijen = []
            else:
                headers = [str(h) if h is not None else "" for h in header_row]
                ontbrekende_kolommen = [kolom for kolom in VERPLICHTE_KOLOMMEN if kolom not in headers]
                if ontbrekende_kolommen:
                    raise ValueError(
                        "Dit lijkt niet het juiste bronbestand.\n"
                        f"Ontbrekende kolommen: {', '.join(ontbrekende_kolommen)}"
                    )
                alle_rijen = [
                    dict(zip(headers, row))
                    for row in ws.iter_rows(min_row=2, values_only=True)
                ]
        finally:
            wb.close()

        bron_bestand = Path(file_path)
        process_data()
        export_button.config(state="normal")
        status_var.set(f"Geladen: {bron_bestand.name}")
        update_preview()
    except Exception as e:
        messagebox.showerror("Fout", str(e))
        status_var.set("Kon het Excel-bestand niet laden.")


def laad_excel_automatisch():
    # Zoekt automatisch een bronbestand naast het script, zodat de gebruiker vaak meteen kan starten.
    script_map = get_app_folder()
    files = sorted(f for f in os.listdir(script_map) if is_bron_excel_bestand(f))
    if not files:
        status_var.set("Geen bronbestand gevonden. Kies handmatig een Excel-bestand.")
        return

    laad_excel_bestand(script_map / files[0])


def kies_excel_bestand():
    # Laat de gebruiker zelf een Excel-bestand kiezen wanneer het automatische bestand niet klopt.
    file_path = filedialog.askopenfilename(
        title="Kies Excel-bestand",
        initialdir=get_app_folder(),
        filetypes=[("Excel-bestanden", "*.xlsx"), ("Alle bestanden", "*.*")],
    )
    if file_path:
        laad_excel_bestand(file_path)


def process_data():
    # Zet ruwe Excel-rijen om naar bruikbare data met gang en locatienummer, en filtert op lege actieve locaties.
    global alle_rijen, gefilterde_rijen

    processed = []
    for row in alle_rijen:
        loc = str(row.get("Location", "") or "").strip()
        if not loc:
            continue

        gang = loc[:3]
        loc_num_str = loc[3:6]
        if not loc_num_str.isnumeric():
            continue

        processed_row = dict(row)
        processed_row.update(
            {
                "Location": loc,
                "Gang": gang,
                "LocatieNummer": int(loc_num_str),
            }
        )
        processed.append(processed_row)

    alle_rijen = processed
    gefilterde_rijen = [
        r
        for r in processed
        if is_valid_location(r)
        and r.get("Location Status") == EMPTY_STATUS
        and is_enabled(r)
    ]
    update_gangen()


def update_gangen():
    # Vult de gang-keuzelijst met alle unieke gangen uit de gefilterde lege locaties.
    unieke_gangen = sorted({r["Gang"] for r in gefilterde_rijen})
    gang_dropdown["values"] = [ALLE_GANGEN] + unieke_gangen
    gang_var.set(ALLE_GANGEN)


def get_location_type(loc):
    # Bepaalt welk uitzonderingsschema bij een locatie hoort op basis van de gangcode.
    gang = str(loc).strip()[:3]
    if gang.startswith("A"):
        return "A"
    if gang in ("BBX", "CCX"):
        return "BBX"
    if gang.startswith("B"):
        try:
            return "BC" if 10 <= int(gang[1:]) <= 21 else "OTHER"
        except ValueError:
            return "OTHER"
    if gang.startswith("C"):
        try:
            return "BC" if 22 <= int(gang[1:]) <= 32 else "OTHER"
        except ValueError:
            return "OTHER"
    return "OTHER"


def is_excluded_location(row):
    # Controleert of het locatienummer voorkomt in de uitzonderingslijst voor dit locatietype.
    location_type = get_location_type(row["Location"])
    return row["LocatieNummer"] in UITGEZONDERDE_LOCATIONS.get(location_type, set())


def houd_lege_locatie(row):
    # Houdt alleen lege locaties over die niet in de uitzonderingslijst staan.
    return not is_excluded_location(row)


def keep_full_juk_location(row):
    # Houdt volle locaties over die juist wel in de uitzonderingslijst staan.
    return is_excluded_location(row)


def get_family(sto_zone):
    # Herkent de familie van een locatie aan de hand van zoekwoorden in Sto_Zone_Cod.
    if not sto_zone:
        return None

    sto_zone_lower = str(sto_zone).lower()
    for keyword, label in zip(FAMILY_KEYWORDS, FAMILY_LABELS):
        if keyword in sto_zone_lower:
            return label
    return None


def get_height(row):
    # Leest de hoogte als getal; bij ontbrekende of ongeldige waarden wordt 0 gebruikt.
    try:
        return float(row.get("Height", 0) or 0)
    except (TypeError, ValueError):
        return 0


def build_family_rows():
    # Bouwt de brondata voor het family-overzicht met alleen lege, actieve locaties onder de maximale hoogte.
    family_rows = []
    for row in alle_rijen:
        if row.get("Location Status") != EMPTY_STATUS:
            continue
        if not is_enabled(row):
            continue

        aisle = str(row.get("Aisle", "") or "").strip()
        if not aisle:
            continue

        height = get_height(row)
        if height >= 1000:
            continue

        family = get_family(row.get("Sto_Zone_Cod", ""))
        if not family:
            continue

        family_rows.append({"Aisle": aisle, "Family": family, "Height": height})
    return family_rows


def apply_borders(ws):
    # Past de standaard dunne rand toe op alle cellen in een werkblad.
    for row in ws.iter_rows():
        for cell in row:
            cell.border = DUNNE_RANDEN


def append_count_pairs(target_row, counts, key_prefix):
    # Voegt per familie de lage en hoge aantallen toe aan een Excel-rij.
    for label in FAMILY_LABELS:
        low_count = counts[(*key_prefix, label, "low")]
        high_count = counts[(*key_prefix, label, "high")]
        target_row += [low_count or "", high_count or ""]


def create_empty_locations_sheet(ws, data_list):
    # Maakt het werkblad met de lijst van lege locaties en een kolom om af te vinken.
    ws.title = "lege locaties"
    ws.append(["Lege Locaties", "Afgevinkt"])
    for row in data_list:
        ws.append([row["Location"], ""])
    apply_borders(ws)


def create_full_locations_sheet(wb, data_full):
    # Maakt het werkblad met volle locaties die op een uitgezonderde derde locatie staan.
    ws = wb.create_sheet(title="pallet in 3e locatie")
    ws.append(["Gang", "Locatie", "Hoogte", "FULL"])
    for row in data_full:
        ws.append(
            [
                row["Gang"],
                row["LocatieNummer"],
                str(row["Location"])[-2:],
                row.get("Location Status", ""),
            ]
        )
    apply_borders(ws)


def create_family_sheet(wb):
    # Maakt het family-overzicht met aantallen per gang, familie en hoogtegroep.
    ws = wb.create_sheet(title="family overzicht")
    family_rows = build_family_rows()
    alle_aisles = sorted({r["Aisle"] for r in family_rows})
    counts = Counter()

    for row in family_rows:
        height_group = "low" if row["Height"] < 232 else "high"
        counts[(row["Aisle"], row["Family"], height_group)] += 1
        counts[(row["Family"], height_group)] += 1

    header = [len(family_rows)]
    for label in FAMILY_LABELS:
        header += [f"{label} (<232)", f"{label} (>=232)"]
    ws.append(header)

    style_family_header(ws)

    for aisle in alle_aisles:
        row_data = [aisle]
        append_count_pairs(row_data, counts, (aisle,))
        ws.append(row_data)

    style_family_body(ws)

    total_row = ["TOTAAL"]
    append_count_pairs(total_row, counts, ())
    ws.append(total_row)
    style_total_row(ws)
    configure_family_sheet_page(ws)


def create_occupancy_sheet(wb):
    # Maakt het extra tabblad met totalen, vrije locaties, bezette locaties en bezettingsgraad per gang.
    ws = wb.create_sheet(title="bezettingsgraad")
    ws.append(["gang", "totaal aantal locaties", "bezet ", "vrij", "bezettingsgraad"])

    bezetting_per_gang = build_occupancy_rows()
    laatste_data_rij = len(bezetting_per_gang) + 2
    ws.append(
        [
            "",
            f"=SUM(B3:B{laatste_data_rij})",
            f"=SUM(C3:C{laatste_data_rij})",
            f"=SUM(D3:D{laatste_data_rij})",
            "=IFERROR(C2/B2,0)",
        ]
    )

    for gang, aantallen in bezetting_per_gang:
        totaal = aantallen[EMPTY_STATUS] + aantallen[FULL_STATUS]
        bezet = aantallen[FULL_STATUS]
        vrij = aantallen[EMPTY_STATUS]
        ws.append([gang, totaal, bezet, vrij, f"=IFERROR(C{ws.max_row + 1}/B{ws.max_row + 1},0)"])

    style_occupancy_sheet(ws, laatste_data_rij)


def build_occupancy_rows():
    # Telt per gang hoeveel ingeschakelde locaties vol en leeg zijn op basis van Enabled? en Location Status.
    counts = {}
    for row in alle_rijen:
        status = row.get("Location Status")
        if status not in (EMPTY_STATUS, FULL_STATUS):
            continue
        if not is_valid_location(row) or not is_enabled(row):
            continue

        gang = row["Gang"]
        if gang not in counts:
            counts[gang] = Counter()
        counts[gang][status] += 1

    return sorted(counts.items())


def style_occupancy_sheet(ws, laatste_data_rij):
    # Maakt het bezettingsgraad-tabblad op zoals het voorbeeldbestand.
    ws.column_dimensions["A"].width = 4.9
    ws.column_dimensions["B"].width = 18.8
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 10.3
    ws.column_dimensions["E"].width = 14.2
    ws.row_dimensions[2].height = 21

    for row in ws.iter_rows(min_row=1, max_row=laatste_data_rij, min_col=1, max_col=5):
        for cell in row:
            cell.border = DUNNE_RANDEN

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cell in ws[2][1:]:
        cell.font = Font(size=16)

    for row in ws.iter_rows(min_row=3, max_row=laatste_data_rij):
        row[0].font = Font(name="Arial", size=10, bold=True)
        row[0].alignment = Alignment(horizontal="center")
        row[4].number_format = "0.0%"

    ws["E2"].number_format = "0.0%"


def style_family_header(ws):
    # Geeft de kopregel van het family-overzicht een donkere achtergrond, witte tekst en randen.
    header_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = DUNNE_RANDEN


def style_family_body(ws):
    # Maakt de datarijen van het family-overzicht op en kleurt de familiekolommen.
    for excel_row in ws.iter_rows(min_row=2):
        for col_idx, cell in enumerate(excel_row):
            cell.border = DUNNE_RANDEN
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)

            if col_idx == 0:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = PatternFill("solid", start_color="ECEFF1", end_color="ECEFF1")
                continue

            family_idx = (col_idx - 1) // 2
            if 0 <= family_idx < len(FAMILY_LABELS):
                color = FAMILY_KLEUREN[FAMILY_LABELS[family_idx]]
                cell.fill = PatternFill("solid", start_color=color, end_color=color)


def style_total_row(ws):
    # Geeft de totaalregel onderaan het family-overzicht dezelfde opvallende stijl als de kopregel.
    total_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    total_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = DUNNE_RANDEN


def configure_family_sheet_page(ws):
    # Stelt kolombreedtes en pagina-instellingen in zodat het family-overzicht netjes op een pagina past.
    ws.column_dimensions["A"].width = 10
    ws.row_dimensions[1].height = 30

    for i in range(len(FAMILY_LABELS) * 2):
        ws.column_dimensions[get_column_letter(i + 2)].width = 11

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.2,
        bottom=0.2,
        header=0.1,
        footer=0.1,
    )


def get_empty_locations_for_selection():
    # Past de huidige gang- en even/oneven-keuze toe en geeft de exporteerbare lege locaties terug.
    gang_selectie = gang_var.get()
    keuze = keuze_var.get()

    if gang_selectie == ALLE_GANGEN:
        data_list = [r for r in gefilterde_rijen if houd_lege_locatie(r)]
    else:
        data_list = [
            r
            for r in gefilterde_rijen
            if r["Gang"] == gang_selectie and houd_lege_locatie(r)
        ]

    if keuze == "Even":
        data_list = [r for r in data_list if r["LocatieNummer"] % 2 == 0]
    elif keuze == "Oneven":
        data_list = [r for r in data_list if r["LocatieNummer"] % 2 == 1]

    data_list.sort(key=lambda r: (r["Gang"], r["LocatieNummer"]))
    return data_list


def get_full_juk_locations():
    # Verzamelt volle locaties die in de speciale uitzonderingsposities vallen.
    data_full = [
        r
        for r in alle_rijen
        if r.get("Location Status") == FULL_STATUS
        and is_valid_location(r)
        and is_enabled(r)
        and keep_full_juk_location(r)
    ]
    data_full.sort(key=lambda r: (r["Gang"], r["LocatieNummer"]))
    return data_full


def update_preview(*_):
    # Toont hoeveel regels er met de huidige filters geëxporteerd worden.
    if not gefilterde_rijen:
        preview_var.set("Nog geen lege locaties geladen.")
        return

    lege_count = len(get_empty_locations_for_selection())
    full_count = len(get_full_juk_locations())
    preview_var.set(
        f"Export bevat {lege_count} lege locaties en {full_count} volle locaties op 3e positie."
    )


def maak_uniek_output_pad(output_path):
    # Maakt een unieke bestandsnaam als er al een exportbestand met dezelfde naam bestaat.
    if not output_path.exists():
        return output_path

    for nummer in range(2, 100):
        kandidaat = output_path.with_name(f"{output_path.stem}_{nummer}{output_path.suffix}")
        if not kandidaat.exists():
            return kandidaat

    return output_path


def filter_and_export():
    # Filtert de ingelezen data op gang en even/oneven keuze, maakt de Excel-werkmap en slaat die op.
    if not gefilterde_rijen:
        messagebox.showinfo("Geen data", "Er zijn nog geen lege locaties geladen om te exporteren.")
        return

    gang_selectie = gang_var.get()
    keuze = keuze_var.get()
    data_list = get_empty_locations_for_selection()
    if not data_list:
        messagebox.showinfo("Geen resultaat", "Er zijn geen lege locaties voor deze selectie.")
        return

    data_full = get_full_juk_locations()

    try:
        g_naam = "Alle_Gangen" if gang_selectie == ALLE_GANGEN else gang_selectie
        bestandsnaam = f"Lege_locaties_{g_naam}_{keuze}.xlsx"
        output_path = maak_uniek_output_pad(get_app_folder() / bestandsnaam)

        wb = Workbook()
        create_empty_locations_sheet(wb.active, data_list)
        create_family_sheet(wb)
        create_occupancy_sheet(wb)
        create_full_locations_sheet(wb, data_full)

        wb.save(output_path)
        os.startfile(output_path)
        status_var.set(f"Bestand opgeslagen en geopend: {output_path.name}")
    except Exception as e:
        messagebox.showerror("Export fout", str(e))


# Bouwt het hoofdvenster van de Tkinter-applicatie.
root = tk.Tk()
root.title("Locatie Tool")
root.geometry("620x320")
root.resizable(False, False)

# Plaatst het hoofdframe met extra ruimte voor labels, status en filterinformatie.
frame = ttk.Frame(root, padding=20)
frame.pack(fill="both", expand=True)

# Maakt Tkinter-variabelen voor de geselecteerde gang, even/oneven keuze en statusregels.
gang_var = tk.StringVar()
keuze_var = tk.StringVar(value="Beide")
status_var = tk.StringVar(value="Bestand laden...")
preview_var = tk.StringVar(value="Nog geen data geladen.")

# Toont een titel en korte context boven de bedieningselementen.
ttk.Label(frame, text="Lege locaties exporteren", font=("Arial", 14, "bold")).grid(
    row=0,
    column=0,
    columnspan=4,
    sticky="w",
    pady=(0, 12),
)

# Maakt een knop waarmee de gebruiker zelf het bronbestand kan kiezen.
ttk.Button(frame, text="Kies Excel-bestand", command=kies_excel_bestand).grid(
    row=1,
    column=0,
    sticky="w",
    pady=(0, 12),
)

# Maakt de keuzelijst voor de gang; de waarden worden gevuld nadat het Excel-bestand is ingelezen.
ttk.Label(frame, text="Gang").grid(row=2, column=0, sticky="w")
gang_dropdown = ttk.Combobox(frame, textvariable=gang_var, width=15, state="readonly")
gang_dropdown.grid(row=3, column=0, sticky="w", padx=(0, 12), pady=(2, 12))
gang_dropdown.bind("<<ComboboxSelected>>", update_preview)

# Maakt de keuzelijst waarmee de gebruiker even, oneven of beide locatienummers kan kiezen.
ttk.Label(frame, text="Locatienummers").grid(row=2, column=1, sticky="w")
keuze_dropdown = ttk.Combobox(
    frame,
    textvariable=keuze_var,
    values=["Even", "Oneven", "Beide"],
    width=10,
    state="readonly",
)
keuze_dropdown.grid(row=3, column=1, sticky="w", padx=(0, 12), pady=(2, 12))
keuze_dropdown.bind("<<ComboboxSelected>>", update_preview)

# Maakt de knop die de gefilterde data exporteert naar een nieuw Excel-bestand.
export_button = ttk.Button(frame, text="Genereer Excel", command=filter_and_export, state="disabled")
export_button.grid(
    row=3,
    column=2,
    sticky="w",
    padx=(0, 12),
    pady=(2, 12),
)

# Toont de laadstatus van het Excel-bestand onder de bedieningselementen.
status_label = ttk.Label(frame, textvariable=status_var)
status_label.grid(row=4, column=0, columnspan=4, sticky="w", pady=(8, 2))

# Toont een live samenvatting van wat er met de huidige filters geëxporteerd wordt.
preview_label = ttk.Label(frame, textvariable=preview_var)
preview_label.grid(row=5, column=0, columnspan=4, sticky="w", pady=(2, 12))

# Geeft extra uitleg zonder dat de gebruiker de code hoeft te kennen.
ttk.Label(
    frame,
    text="Tip: outputbestanden krijgen automatisch een volgnummer als de naam al bestaat.",
    foreground="#555555",
).grid(row=6, column=0, columnspan=4, sticky="w")

# Start kort na het openen automatisch het inlezen van Excel en daarna de Tkinter-eventloop.
root.after(50, laad_excel_automatisch)
root.mainloop()
