import os
import sys
import tkinter as tk
from collections import Counter
from pathlib import Path
from tkinter import messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


ALL_GANGEN = "Alle Gangen"
EMPTY_STATUS = "E - Empty"
FULL_STATUS = "F - Full"
VALID_LOCATION_PREFIXES = ("A", "B", "C")

EXCLUDE_LOCATIONS = {
    "A": set(range(9, 211, 8)) | set(range(10, 211, 8)),
    "BBX": set(range(3, 80, 4)),
    "BC": set(range(5, 199, 8)) | set(range(6, 199, 8)),
}

FAMILY_KEYWORDS = ("big", "med-lng", "med-shrt", "sml-lng", "sml-shrt", "ground")
FAMILY_LABELS = ("Big", "Med-Lng", "Med-Shrt", "Sml-Lng", "Sml-Shrt", "Ground")
FAMILY_COLORS = {
    "Big": "D9EAD3",
    "Med-Lng": "CFE2F3",
    "Med-Shrt": "EAD1DC",
    "Sml-Lng": "FFF2CC",
    "Sml-Shrt": "F4CCCC",
    "Ground": "E6E0EC",
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

rows_raw = []       # Alle rijen als lijst van dicts
rows_filtered = []  # Gefilterde rijen (A/B/C + Empty)


def get_app_folder():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def is_enabled(row):
    return row.get("Enabled?") == 1


def is_valid_location(row):
    location = str(row.get("Location", "") or "")
    return location.startswith(VALID_LOCATION_PREFIXES)


def load_excel_auto():
    global rows_raw

    script_map = get_app_folder()
    try:
        files = sorted(f for f in os.listdir(script_map) if f.endswith(".xlsx"))
        if not files:
            return

        file_path = script_map / files[0]
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(values_only=True), None)
            if header_row is None:
                rows_raw = []
            else:
                headers = [str(h) if h is not None else "" for h in header_row]
                rows_raw = [
                    dict(zip(headers, row))
                    for row in ws.iter_rows(min_row=2, values_only=True)
                ]
        finally:
            wb.close()

        process_data()
        status_label.config(text=f"Loaded: {files[0]}")
    except Exception as e:
        messagebox.showerror("Fout", str(e))


def process_data():
    global rows_raw, rows_filtered

    processed = []
    for row in rows_raw:
        loc = str(row.get("Location", "") or "").strip()
        if not loc:
            continue

        gang = loc[:3]
        loc_num_str = loc[3:6]
        if not loc_num_str.isnumeric():
            continue

        processed_row = dict(row)
        processed_row.update(
            {
                "Location": loc,
                "Gang": gang,
                "LocatieNummer": int(loc_num_str),
            }
        )
        processed.append(processed_row)

    rows_raw = processed
    rows_filtered = [
        r
        for r in processed
        if is_valid_location(r)
        and r.get("Location Status") == EMPTY_STATUS
        and is_enabled(r)
    ]
    update_gangen()


def update_gangen():
    unieke_gangen = sorted({r["Gang"] for r in rows_filtered})
    gang_dropdown["values"] = [ALL_GANGEN] + unieke_gangen
    gang_var.set(ALL_GANGEN)


def get_location_type(loc):
    gang = str(loc).strip()[:3]
    if gang.startswith("A"):
        return "A"
    if gang in ("BBX", "CCX"):
        return "BBX"
    if gang.startswith("B"):
        try:
            return "BC" if 10 <= int(gang[1:]) <= 21 else "OTHER"
        except ValueError:
            return "OTHER"
    if gang.startswith("C"):
        try:
            return "BC" if 22 <= int(gang[1:]) <= 32 else "OTHER"
        except ValueError:
            return "OTHER"
    return "OTHER"


def is_excluded_location(row):
    location_type = get_location_type(row["Location"])
    return row["LocatieNummer"] in EXCLUDE_LOCATIONS.get(location_type, set())


def keep_empty_location(row):
    return not is_excluded_location(row)


def keep_full_juk_location(row):
    return is_excluded_location(row)


def get_family(sto_zone):
    if not sto_zone:
        return None

    sto_zone_lower = str(sto_zone).lower()
    for keyword, label in zip(FAMILY_KEYWORDS, FAMILY_LABELS):
        if keyword in sto_zone_lower:
            return label
    return None


def get_height(row):
    try:
        return float(row.get("Height", 0) or 0)
    except (TypeError, ValueError):
        return 0


def build_family_rows():
    family_rows = []
    for row in rows_raw:
        if row.get("Location Status") != EMPTY_STATUS:
            continue
        if not is_enabled(row):
            continue

        aisle = str(row.get("Aisle", "") or "").strip()
        if not aisle:
            continue

        height = get_height(row)
        if height >= 1000:
            continue

        family = get_family(row.get("Sto_Zone_Cod", ""))
        if not family:
            continue

        family_rows.append({"Aisle": aisle, "Family": family, "Height": height})
    return family_rows


def apply_borders(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER


def append_count_pairs(target_row, counts, key_prefix):
    for label in FAMILY_LABELS:
        low_count = counts[(*key_prefix, label, "low")]
        high_count = counts[(*key_prefix, label, "high")]
        target_row += [low_count or "", high_count or ""]


def create_empty_locations_sheet(ws, data_list):
    ws.title = "lege locaties"
    ws.append(["Lege Locaties", "Afgevinkt"])
    for row in data_list:
        ws.append([row["Location"], ""])
    apply_borders(ws)


def create_full_locations_sheet(wb, data_full):
    ws = wb.create_sheet(title="pallet in 3e locatie")
    ws.append(["Gang", "Locatie", "Hoogte", "FULL"])
    for row in data_full:
        ws.append(
            [
                row["Gang"],
                row["LocatieNummer"],
                str(row["Location"])[-2:],
                row.get("Location Status", ""),
            ]
        )
    apply_borders(ws)


def create_family_sheet(wb):
    ws = wb.create_sheet(title="family overzicht")
    family_rows = build_family_rows()
    alle_aisles = sorted({r["Aisle"] for r in family_rows})
    counts = Counter()

    for row in family_rows:
        height_group = "low" if row["Height"] < 232 else "high"
        counts[(row["Aisle"], row["Family"], height_group)] += 1
        counts[(row["Family"], height_group)] += 1

    header = [len(family_rows)]
    for label in FAMILY_LABELS:
        header += [f"{label} (<232)", f"{label} (>=232)"]
    ws.append(header)

    style_family_header(ws)

    for aisle in alle_aisles:
        row_data = [aisle]
        append_count_pairs(row_data, counts, (aisle,))
        ws.append(row_data)

    style_family_body(ws)

    total_row = ["TOTAAL"]
    append_count_pairs(total_row, counts, ())
    ws.append(total_row)
    style_total_row(ws)
    configure_family_sheet_page(ws)


def style_family_header(ws):
    header_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_family_body(ws):
    for excel_row in ws.iter_rows(min_row=2):
        for col_idx, cell in enumerate(excel_row):
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)

            if col_idx == 0:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = PatternFill("solid", start_color="ECEFF1", end_color="ECEFF1")
                continue

            family_idx = (col_idx - 1) // 2
            if 0 <= family_idx < len(FAMILY_LABELS):
                color = FAMILY_COLORS[FAMILY_LABELS[family_idx]]
                cell.fill = PatternFill("solid", start_color=color, end_color=color)


def style_total_row(ws):
    total_fill = PatternFill("solid", start_color="37474F", end_color="37474F")
    total_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def configure_family_sheet_page(ws):
    ws.column_dimensions["A"].width = 10
    ws.row_dimensions[1].height = 30

    for i in range(len(FAMILY_LABELS) * 2):
        ws.column_dimensions[get_column_letter(i + 2)].width = 11

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.2,
        bottom=0.2,
        header=0.1,
        footer=0.1,
    )


def filter_and_export():
    if not rows_filtered:
        return

    gang_selectie = gang_var.get()
    keuze = keuze_var.get()

    if gang_selectie == ALL_GANGEN:
        data_list = [r for r in rows_filtered if keep_empty_location(r)]
    else:
        data_list = [
            r
            for r in rows_filtered
            if r["Gang"] == gang_selectie and keep_empty_location(r)
        ]

    if keuze == "Even":
        data_list = [r for r in data_list if r["LocatieNummer"] % 2 == 0]
    elif keuze == "Oneven":
        data_list = [r for r in data_list if r["LocatieNummer"] % 2 == 1]

    data_list.sort(key=lambda r: (r["Gang"], r["LocatieNummer"]))

    data_full = [
        r
        for r in rows_raw
        if r.get("Location Status") == FULL_STATUS
        and is_valid_location(r)
        and is_enabled(r)
        and keep_full_juk_location(r)
    ]
    data_full.sort(key=lambda r: (r["Gang"], r["LocatieNummer"]))

    try:
        g_naam = "Alle_Gangen" if gang_selectie == ALL_GANGEN else gang_selectie
        bestandsnaam = f"Lege_locaties_{g_naam}_{keuze}.xlsx"
        output_path = get_app_folder() / bestandsnaam

        wb = Workbook()
        create_empty_locations_sheet(wb.active, data_list)
        create_family_sheet(wb)
        create_full_locations_sheet(wb, data_full)

        wb.save(output_path)
        messagebox.showinfo("Succes", f"Bestand opgeslagen:\n{bestandsnaam}")
    except Exception as e:
        messagebox.showerror("Export fout", str(e))


root = tk.Tk()
root.title("Locatie Tool")
root.geometry("500x250")

frame = tk.Frame(root)
frame.pack(pady=20)

gang_var = tk.StringVar()
keuze_var = tk.StringVar(value="Beide")

gang_dropdown = ttk.Combobox(frame, textvariable=gang_var, width=15, state="readonly")
gang_dropdown.grid(row=0, column=0, padx=5)

keuze_dropdown = ttk.Combobox(
    frame,
    textvariable=keuze_var,
    values=["Even", "Oneven", "Beide"],
    width=10,
    state="readonly",
)
keuze_dropdown.grid(row=0, column=1, padx=5)

tk.Button(frame, text="Genereer Excel", command=filter_and_export).grid(
    row=0,
    column=2,
    padx=5,
)

status_label = tk.Label(root, text="Bestand laden...")
status_label.pack(pady=10)

root.after(100, load_excel_auto)
root.mainloop()
